
# html

#from HTMLParser import HTMLParser
#
## create a subclass and override the handler methods
#class MyHTMLParser(HTMLParser):
#    def handle_starttag(self, tag, attrs):
#        print "Encountered a start tag:", tag
#
#    def handle_endtag(self, tag):
#        print "Encountered an end tag :", tag
#
#    def handle_data(self, data):
#        print "Encountered some data  :", data
#
## instantiate the parser and fed it some HTML
#parser = MyHTMLParser()
#parser.feed('<html><head><title>Test</title></head>'
#            '<body><h1>Parse me!</h1></body></html>')

#from lxml import html
#import requests
#
#page = requests.get('http://www.pealim.com/dict/1')
#tree = html.fromstring(page.content)

#This will create a list of buyers:
##buyers = tree.xpath('//div[@title="buyer-name"]/text()')
#This will create a list of prices
#transcription = tree.xpath('//div[@class="transcription"]/text()')
#menukad = tree.xpath('//span[@class="menukad"]/text()')




#######################""
#xlwt

#
#import xlwt
#from datetime import datetime
#
#font0 = xlwt.Font()
#font0.name = 'Times New Roman'
#font0.colour_index = 2
#font0.bold = True
#
#style0 = xlwt.XFStyle()
#style0.font = font0
#
#style1 = xlwt.XFStyle()
#style1.num_format_str = 'D-MMM-YY'
#
#wb = xlwt.Workbook()
#ws = wb.add_sheet('A Test Sheet')
#
#ws.write(0, 0, 'Test', style0)
#ws.write(1, 0, datetime.now(), style1)
#ws.write(2, 0, 1)
#ws.write(2, 1, 1)
#ws.write(2, 2, xlwt.Formula("A3+B3"))
#
#wb.save('example.xls')


##
#
#import xlsxwriter
#
#
## Create an new Excel file and add a worksheet.
#workbook = xlsxwriter.Workbook('demo.xlsx')
#worksheet = workbook.add_worksheet()
#
## Widen the first column to make the text clearer.
#worksheet.set_column('A:A', 20)
#
## Add a bold format to use to highlight cells.
#bold = workbook.add_format({'bold': True})
#
## Write some simple text.
#worksheet.write('A1', 'Hello')
#
## Text with formatting.
#worksheet.write('A2', 'World', bold)
#
## Write some numbers, with row/column notation.
#worksheet.write(2, 0, 123)
#worksheet.write(3, 0, 123.456)
#
## Insert an image.
#worksheet.insert_image('B5', 'logo.png')
#
#workbook.close()


            
            







#

#import xlwt

#
#x=1
#y=2
#z=3
#
#list1=[2.34,4.346,4.234]

#book = xlwt.Workbook(encoding="utf-8")

#sheet1 = book.add_sheet("Sheet 1")

#sheet1.write(0, 0, "Display")
#sheet1.write(1, 0, "Dominance")
#sheet1.write(2, 0, "Test")
#
#sheet1.write(0, 1, x)
#sheet1.write(1, 1, y)
#sheet1.write(2, 1, z)
#
#sheet1.write(4, 0, "Stimulus Time")
#sheet1.write(4, 1, "Reaction Time")

#i=4
#
#for n in list1:
#    i = i+1
#    sheet1.write(i, 0, n)

#transcription = tree.xpath('//div[@class="transcription"]/text()')
#menukad = tree.xpath('//span[@class="menukad"]/text()')



####################

#from lxml import html
#from lxml import etree
#import requests
#import xlwt


#book = xlwt.Workbook(encoding="utf-8")
#sheet1 = book.add_sheet("Sheet1")
#b=0
#a=1
#page = requests.get('http://www.pealim.com/dict/'+str(a))
#for a in range(1,500):
#    try:
#    page = requests.get('http://www.pealim.com/dict/'+str(a))
#    if page.content != '':
#        tree = html.fromstring(page.content)
        #transcription = tree.xpath('//div[@class="transcription"]') #/text()
#        lead = tree.xpath('//div[@class="lead"]/text()')    
#        menukad = tree.xpath('//span[@class="menukad"]/text()')
#        if menukad!=[]:
#            sheet1.write(0, b, a)
#            sheet1.write(1, b, lead[0])
#            i=2
#            for x in menukad:
#                sheet1.write(i, b, x)
#                i=i+1
#            b=b+1
#            print(str(a)+" : ok")
#        else:
#            print(str(a)+" : echec")
#    else:
#        print(str(a)+" : echec de type 2")

#    except:
#        print(str(a)+" : echec");
#        pass

#except:
#    print "Unexpected error"
#    book.save("trial6.xls")
#    raise

#cd "C:\\Users\\Nicolas\\Desktop\\"
#book.save('a1.xls')


###########################

#from lxml import html
#import requests
#import xlwt
#
#
#book = xlwt.Workbook(encoding="utf-8")
#sheet1 = book.add_sheet("Sheet1")
#b=0
#a=1
#page = requests.get('http://www.pealim.com/dict/'+str(a))
#for a in range(1,500):
#    page = requests.get('http://www.pealim.com/dict/'+str(a))
#    if page.content != '':
#        tree = html.fromstring(page.content)
#        lead = tree.xpath('//div[@class="lead"]/text()')    
#        menukad = tree.xpath('//span[@class="menukad"]/text()')
#        if menukad!=[]:
#            sheet1.write(0, b, a)
#            sheet1.write(1, b, lead[0])
#            i=2
#            for x in menukad:
#                sheet1.write(i, b, x)
#                i=i+1
#            b=b+1
#            print(str(a)+" : ok")
#        else:
#            print(str(a)+" : echec")
#    else:
#        print(str(a)+" : echec de type 2")
#book.save('a1.xls')




############################

#from lxml import html
#import requests
#import csv
#
#sheet1 = csv.writer(open('file.csv', 'wb'))
#
#b=0
#a=1
#page = requests.get('http://www.pealim.com/dict/'+str(a))
#for a in range(1,500):
#    page = requests.get('http://www.pealim.com/dict/'+str(a))
#    if page.content != '':
#        tree = html.fromstring(page.content)
#        lead = tree.xpath('//div[@class="lead"]/text()')    
#        menukad = tree.xpath('//span[@class="menukad"]/text()')
#        if menukad!=[]:
#            sheet1.writerow([str(a),lead[0]]+menukad)
#            print(str(a)+" : ok")
#        else:
#            print(str(a)+" : echec")
#    else:
#        print(str(a)+" : echec de type 2")



###############################
#
#from openpyxl import Workbook
#wb = Workbook()
#ws = wb.active
#ws['A1'] = 42
#ws.append([1, 2, 3])
#import datetime
#ws['A2'] = datetime.datetime.now()
#wb.save("sample.xlsx")
#
##############################################

#from openpyxl import Workbook
#from lxml import html
#import requests
#
#wb = Workbook()
#sheet1 = wb.active
#
#for a in range(1,2809):
#    page = requests.get('http://www.pealim.com/dict/'+str(a))
#    if page.content != '':
#        tree = html.fromstring(page.content)
#        lead = tree.xpath('//div[@class="lead"]/text()')
#        menukad = tree.xpath('//span[@class="menukad"]/text()')
#        binyan = tree.xpath('/html/body/div/div[1]/p[1]/b/text()')
#        if menukad!=[]:
#            sheet1.append([str(a),lead[0],binyan[0]]+menukad)
#            print(str(a)+" : ok")
#        else:
#            print(str(a)+" : echec")
#    else:
#        print(str(a)+" : echec de type 2")
#
#cd "C:\\Users\\Nicolas\\Desktop\\"
#pwd
#wb.save("sample.xlsx")

#2644 premier pas normal
#2809 last normal, 2810 premier chelou
# fin 3061



###########################
#from openpyxl import Workbook
#from openpyxl import writer


#####################################
# VERBS
################################

import openpyxl
from lxml import html
import requests

wb = openpyxl.Workbook()
sheet1=wb.active
sheet2=wb.create_sheet()

ll=["INF-L","AP-ms","AP-fs","AP-mp","AP-fp","PERF-1s",
"PERF-2ms","PERF-2fs","PERF-3ms","PERF-3fs","PERF-1p",
"PERF-2mp","PERF-3p","IMPF-1s","IMPF-2ms","IMPF-2fs",
"IMPF-3ms","IMPF-1p","IMPF-2mp","IMPF-3mp","IMP-2ms","IMP-2fs","IMP-2mp"]

#2809
for a in range(2644,3061):
    page = requests.get('http://www.pealim.com/dict/'+str(a))
    
    if page.content != '':
        tree = html.fromstring(page.content)
        lead = tree.xpath('//div[@class="lead"]/text()')
        if lead!=['The page you are looking for cannot be found.']:
            typ=tree.xpath('/html/body/div/div[1]/p[1]/text()')[0]
            if typ[0:4]==u'Verb':
                translation=lead[0]
                binyan = tree.xpath('/html/body/div/div[1]/p[1]/b/text()')[0]
                menukad=[]
                francais=[]
                for b in ll:
                    txt1=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/text()[1]")
                    if(txt1!=[]):
                        txt1=txt1[0]
                        txt2=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/b/text()")[0]
                        txt3_temp=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/text()[2]")
                        if txt3_temp==[]:
                            txt3=''
                        else:
                            txt3=txt3_temp[0]
                        francais.append(txt1+txt2+txt3)
                        menukad.append(
                          tree.xpath(
                            "//*[@id=\""+b+"\"]/div[1]/div[1]/span[@class=\"menukad\"]/text()")[0])
                    
                roots=tree.xpath("/html/body/div/div[1]/p[2]/span/text()")[0]
                roots_sep=roots.rsplit(' - ')
                while len(roots_sep)<4:
                    roots_sep.append('')
                roots_sep.reverse()
                
                # num, meaning, binyan, root letters, conjugation
                aaa=[str(a),translation,binyan]+roots_sep
                sheet1.append(aaa+menukad)
                sheet2.append(aaa+francais)
                print(str(a)+" : ok")
            else:
                print(str(a)+" : not a verb")
        else:
            print(str(a)+" : echec")
    else:
        print(str(a)+" : echec de type 2")


cd "C:\\Users\\Nicolas\\Desktop\\"
pwd
wb.save("Pealim_verbs.xlsx")

#######################"
#.("sample.xlsx")

#//*[@id="IMPF-3fp"]/div[1]/div[2]/text()[1]
#//*[@id="IMP-2fp"]/div[1]/div[2]/b
#//*[@id="IMPF-3fp"]/div[1]/div[2]/text()[2]

#//*[@id="IMPF-3fp"]/div[1]/div[1]/span
#ou
#//*[@id="IMPF-3fp"]/div[1]/div[1]/span[1]
# ----->
#//*[@id="IMPF-3fp"]/div[1]/div[1]/span[@class="menukad"]

# racine, sens, groupe

#########################"

#//*[@id="s"]/div[1]/div[2]/text()[1]
#//*[@id="s"]/div[1]/div[2]/b/text()
#//*[@id="s"]/div[1]/div[2]/text()[2]
#//*[@id="s"]/div[1]/div[1]/span/text()

#//*[@id="p"]/div[1]/div[2]/text()[1]

################################"
# NOUN  -----> TO FINISH
############################

import openpyxl
from lxml import html
import requests

wb = openpyxl.Workbook()
sheet1=wb.active
sheet2=wb.create_sheet()

ll=["s","p"]

#2809
for a in range(2664,3061):
    page = requests.get('http://www.pealim.com/dict/'+str(a))
    
    if page.content != '':
        tree = html.fromstring(page.content)
        lead = tree.xpath('//div[@class="lead"]/text()')
        if lead!=['The page you are looking for cannot be found.']:
            typ=tree.xpath('/html/body/div/div[1]/p[1]/text()')[0]
            if typ[0:4]==u'Noun':
                translation=lead[0]
                binyan = tree.xpath('/html/body/div/div[1]/p[1]/b/text()')
                if binyan!=[]:
                    binyan=binyan[0]
                else:
                    binyan=''
                menukad=[]
                francais=[]
                for b in ll:
                    txt1=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/text()[1]")
                    if(txt1!=[]):
                        txt1=txt1[0]
                        txt2=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/b/text()")[0]
                        txt3_temp=tree.xpath("//*[@id=\""+b+"\"]/div[1]/div[2]/text()[2]")
                        if txt3_temp==[]:
                            txt3=''
                        else:
                            txt3=txt3_temp[0]
                        francais.append(txt1+txt2+txt3)
                        menukad.append(
                          tree.xpath(
                            "//*[@id=\""+b+"\"]/div[1]/div[1]/span[@class=\"menukad\"]/text()")[0])
                    
                roots=tree.xpath("/html/body/div/div[1]/p[2]/span/text()")
                if roots==[]:
                    roots=''
                else:
                    roots=roots[0]
                roots_sep=roots.rsplit(' - ')
                while len(roots_sep)<4:
                    roots_sep.append('')
                roots_sep.reverse()
                
                # num, meaning, binyan, root letters, conjugation
                aaa=[str(a),translation,binyan]+roots_sep
                sheet1.append(aaa+menukad)
                sheet2.append(aaa+francais)
                print(str(a)+" : ok")
            else:
                print(str(a)+" : not a noun")
        else:
            print(str(a)+" : echec")
    else:
        print(str(a)+" : echec de type 2")


cd "C:\\Users\\Nicolas\\Desktop\\"
pwd
wb.save("Pealim_nouns.xlsx")




