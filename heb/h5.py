#conj.xlsx
#fr2heb_2
#cube1
#EDCB 4321
#FGHIJ 56789


import unicodedata
import openpyxl

#from openpyxl.utils import (_get_column_letter)
alphabet=[u'\u05D0',u'\u05D1',u'\u05D2',u'\u05D3',
          u'\u05D4',u'\u05D5',u'\u05D6',u'\u05D7',
          u'\u05D8',u'\u05D9',u'\u05DB',
          u'\u05DC',u'\u05DE',
          u'\u05E0',u'\u05E1',u'\u05E2',
          u'\u05E4',u'\u05E6',u'\u05E7',
          u'\u05E8',u'\u05E9',u'\u05EA']
#unichr, func1
#ord(u'\u0010')
#u'\x10'
#u'\231'
#a=u'\u05D0' # --> aleph
#ord(a)   # --> 1488
#unichr(40960)
a=u'\u05DA' cafSofit  u'\u05DA' caf
a= u'\u05DD' memSofit   u'\u05DE' mem
a=u'\u05DF' nunSofit u'\u05E0' nun
a=u'\u05E3' peSofit  u'\u05E4' pe
a=u'\u05E5'  tsSofit   a=u'\u05E6'  ts



def func1(a):
    if a<=11:
        return a
    elif a<=14:#caf
        return a-1
    elif a<=16:#mem
        return a-2
    elif a<=20:#noun
        return a-3
    elif a<=22:#pe
        return a-4
    else:#tsadi
        return a-5

cd 'C:/Users/Nicolas/Desktop/'         

wb = openpyxl.load_workbook('Racines et groupes.xlsx') #'conj.xlsx')
#wb.get_sheet_names()
s1 = wb.active
#s1 = wb.get_sheet_by_name('fr2heb_2')

wb_out=openpyxl.Workbook()


for k in range(1,6):
    print k
    #s2 = wb.get_sheet_by_name('cube'+str(k))
    s2=wb_out.create_sheet(title='cube'+str(k))
    for i in range(0,22):
        car=alphabet[i]
        s2.cell(row=1,column=i+2).value=car
        s2.cell(row=i+2,column=1).value=car
    for row in range(2, 1518): #s1.max_row  255
        print('  '+str(row))
        if s1.cell(row=row,column=4+k).value != None: #5+k
            #EDCB
            lettre1 = s1['D' + str(row)].value
            lettre2 = s1['C' + str(row)].value
            lettre3 = s1['B' + str(row)].value
            lettre4 = s1['A' + str(row)].value

            if lettre4!=None:
                lettre2=lettre2+lettre3
                lettre3=lettre4
                        
            num1=func1(ord(lettre1[0])-1487)+1
            num3=func1(ord(lettre3[0])-1487)+1

            a=s2.cell(row=num3,column=num1).value
            #column_letter = _get_column_letter(num1+1)
            if(a==None):
                s2.cell(row=num3,column=num1).value=lettre2
                #s2[column_letter+str(num3+1)]=lettre2
            else:
                s2.cell(row=num3,column=num1).value=a+', '+lettre2
                #s2[column_letter+str(num3+1)]=a+', '+lettre2

l=wb_out.get_sheet_names()
nom1=l[0]
wb_out.remove_sheet(wb_out.get_sheet_by_name(nom1.decode('utf8')))

wb_out.save('conj222.xlsx')




