# 428

import openpyxl

cd 'C:/Users/Nicolas/Desktop/'

wb1 = openpyxl.load_workbook('Pealim_verbs.xlsx')
s1= wb1.get_sheet_by_name('Sheet3')

wb2 = openpyxl.Workbook()
s2=wb2.active


groupes=[u"PA'AL",u"PI'EL",u"HIF'IL",
        u"HITPA'EL",u"NIF'AL"]

j1=0
i2=0
l=['','','','']
l_old=l[:]

while True:
    j1=j1+1
    
    groupe=s1.cell(row=3,column=j1).value    
    anglais= s1.cell(row=2,column=j1).value
    if anglais==None:
        break
        
    try:
        col=5+groupes.index(groupe)
    except:
        continue
    
    l[0]= s1.cell(row=7,column=j1).value
    l[1]= s1.cell(row=6,column=j1).value
    l[2]= s1.cell(row=5,column=j1).value
    l[3]= s1.cell(row=4,column=j1).value    
    if l != l_old:
        i2=i2+1
        s2.cell(row=i2,column=1).value=l[3]
        s2.cell(row=i2,column=2).value=l[2]
        s2.cell(row=i2,column=3).value=l[1]
        s2.cell(row=i2,column=4).value=l[0]
        l_old=l[:]
    print(str(j1)+'\t'+str(i2))
    
    s2.cell(row=i2,column=col).value=anglais


wb2.save('machin1.xlsx')














