

cd 'C:/Users/Nicolas/Desktop/'

import openpyxl

wb = openpyxl.load_workbook('coco.xlsx')
sheet= wb.worksheets[0]
row_count=sheet.max_row
column_count=sheet.max_column


aaa=0
for ligne in range(2,row_count+1):
    aaa=aaa+1
    print aaa
    for t in [1,2]:
        a=sheet.cell(row=ligne,column=t).value
        if a is not None :
            a=a.replace(u'\u05DA',u'\u05DB')
            a=a.replace(u'\u05DD', u'\u05DE')
            a=a.replace(u'\u05DF' , u'\u05E0')
            a=a.replace(u'\u05E3',  u'\u05E4')
            a=a.replace(u'\u05E5',  u'\u05E6')
            sheet.cell(row=ligne,column=t).value=a


#for onglet in range(1,38):
#    sheet= wb.worksheets[onglet]
#    row_count=sheet.max_row
#    print row_count
 
wb.save('v44.xlsx')

# 16 17 (piel 4)
# 28 29 (hitpael 4)

# 8-15    h-o
# 4 ou 5     d ou e
# 31 onglets en tout (1er inutile)
# 450 lignes



#________________________




#unichr, func1
#ord(u'\u0010')
#u'\x10'
#u'\231'
#a=u'\u05D0' # --> aleph
#ord(a)   # --> 1488
#unichr(40960)


#a=u'\u05DA' cafSofit  u'\u05DB' caf
#a= u'\u05DD' memSofit   u'\u05DE' mem
#a=u'\u05DF' nunSofit u'\u05E0' nun
#a=u'\u05E3' peSofit  u'\u05E4' pe
#a=u'\u05E5'  tsSofit   a=u'\u05E6'  ts


#import unicodedata
#u = unichr(233) + unichr(0x0bf2) + unichr(3972) + unichr(6000) + unichr(13231)
#for i, c in enumerate(u):
#    print i, '%04x' % ord(c), unicodedata.category(c),
#    print unicodedata.name(c)

#s.replace('feather', 'sand')