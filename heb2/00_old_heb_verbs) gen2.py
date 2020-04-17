
#http://openpyxl.readthedocs.io/en/default/styles.html

cd 'C:\Users\Nicolas\Desktop\interets\0) Langues\0) docs generaux'

import openpyxl
wb1 = openpyxl.load_workbook('2)tableau_hi.xlsx')
sheet1=wb1.worksheets[0]
row_count=sheet1.max_row
column_count=sheet1.max_column
for a in range(2,3):
    for b in range(2,10):
        print sheet1.cell(row=b,column=a).value



wb1.save('testDisplay.xlsx')


#wb2=openpyxl.Workbook()
#sheet2=wb2.active
#sheet2.title = 'txt'
#wb2.create_sheet('pron') 
#for sheet in wb2.worksheets:
#    print sheet.title
#sheet3=wb2.worksheets[1]






