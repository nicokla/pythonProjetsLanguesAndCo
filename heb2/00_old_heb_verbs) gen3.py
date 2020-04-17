

cd 'C:\Users\Nicolas\Desktop\interets\0) Langues\0) docs generaux\7_dialectes arabes du levant'

import openpyxl

wb1 = openpyxl.load_workbook('ar_pal.xlsx')
#wb1 = openpyxl.load_workbook('ar_pal_phrases.xlsx')

row_count=sheet1.max_row
column_count=sheet1.max_column
wb1.create_sheet('new')  #sheet1=wb1.worksheets[0]
nbOnglets=len(wb1.worksheets)

sheet1=wb1.worksheets[1]
sheet2=wb1.worksheets[nbOnglets-1] #newly created, last one

for a in range(2,3):
    for b in range(2,10):
        print sheet1.cell(row=b,column=a).value

a0=2
a1=13
b0=2
b1=95
a2=a1-a0+1
b2=b1-b0+1
c=a2*b2
i=2
for a in range(a0,a1+1):
    for b in range(b0,b1+1):
        #sheet2.cell(row=i,column=1).value=sheet1.cell(row=b,column=a).value
        sheet2.cell(row=b,column=a).value=sheet1.cell(row=i,column=2).value
        i+=1

    
    
wb1.save('test.xlsx')
