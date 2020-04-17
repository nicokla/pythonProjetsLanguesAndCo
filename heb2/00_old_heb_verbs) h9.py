

#cd 'C:/Users/Nicolas/Desktop/'
cd 'C:\Users\Nicolas\Desktop\all\Documents\Pas technologique\0) Langues\01) hebreu\0) docs excel et words\voc'
import openpyxl

wb = openpyxl.load_workbook('10000heb_range_1.xlsx')
sheet= wb.worksheets[0]
row_count=sheet.max_row
column_count=sheet.max_column


for r in range(1,row_count+1):
    a=sheet.cell(row=r,column=2).value
    if a is not None :
        temp=a
    else:
        sheet.cell(row=r,column=2).value=temp
    #print r, temp


for r in range(1,row_count+1):
    a=sheet.cell(row=r,column=2).value
    if a is not None :
        temp=a
    else:
        sheet.cell(row=r,column=2).value=temp
    s=len(temp)
    sheet.cell(row=r,column=5).value=s
    for i in range(s):
        sheet.cell(row=r,column=i+6).value=temp[i]



#for onglet in range(1,38):
#    sheet= wb.worksheets[onglet]
#    row_count=sheet.max_row
#    print row_count
 
wb.save('coucou2.xlsx')


#####################

from collections import Counter
import numpy as np
import matplotlib.pyplot as plt


a = ['a', 'a', 'a', 'a', 'b', 'b', 'c', 'c', 'c', 'd', 'e', 'e', 'e', 'e', 'e']
#letter_counts.keys()

def plot_bar_from_counter(counter, ax=None):
    if ax is None:
        fig = plt.figure()
        ax = fig.add_subplot(111)

    frequencies = counter.values()
    names = counter.keys()

    x_coordinates = np.arange(len(counter))
    ax.bar(x_coordinates, frequencies, align='center')

    ax.xaxis.set_major_locator(plt.FixedLocator(x_coordinates))
    ax.xaxis.set_major_formatter(plt.FixedFormatter(names))
    
    return ax


numlet=5
sheet= wb.worksheets[numlet-2]
row_count=sheet.max_row
column_count=sheet.max_column

let=4
l = []
for col in sheet.columns[4+let]:
    l.append(col.value)
letter_counts = Counter(l)
plot_bar_from_counter(letter_counts)
plt.show()
liste=letter_counts.keys()
for tt in liste:
    print tt


# 16 17 (piel 4)
# 28 29 (hitpael 4)

# 8-15    h-o
# 4 ou 5     d ou e
# 31 onglets en tout (1er inutile)
# 450 lignes



#________________________




#unichr, func1
#ord(u'\u0010')
#u'\x10'
#u'\231'
#a=u'\u05D0' # --> aleph
#ord(a)   # --> 1488
#unichr(40960)


#a=u'\u05DA' cafSofit  u'\u05DB' caf
#a= u'\u05DD' memSofit   u'\u05DE' mem
#a=u'\u05DF' nunSofit u'\u05E0' nun
#a=u'\u05E3' peSofit  u'\u05E4' pe
#a=u'\u05E5'  tsSofit   a=u'\u05E6'  ts


#import unicodedata
#u = unichr(233) + unichr(0x0bf2) + unichr(3972) + unichr(6000) + unichr(13231)
#for i, c in enumerate(u):
#    print i, '%04x' % ord(c), unicodedata.category(c),
#    print unicodedata.name(c)

#s.replace('feather', 'sand')