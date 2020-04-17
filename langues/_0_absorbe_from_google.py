

#####################

import time
import os
#import copy
import sys

from googletrans import Translator
translator = Translator()

import openpyxl

from openpyxl.utils import coordinate_from_string, column_index_from_string



#############################

langues1 = ['fr','iw','ar','ru','es',
           'zh-CN','hi','el','tr',
           'de','ro','yi','sw','ja','pt','it']
fileNames1=['100) eu_lat) French',
'200) sem) Hebrew','200) sem) Arabic',
'120) eu_sla-e) Russian','100) eu_lat) Spanish',
'520) as_1) Chinese','510) as_in-n) Hindi',
'160) eu_uncl) Greek','300) tur) Turkish',
'110) eu_ger-w) German','100) eu_lat) Romanian',
'110) eu_ger-w) Yiddish','400) af) Swahili',
'550) as_3) Japanese','100) eu_lat) Portuguese',
'100) eu_lat) Italian']

langues2=['nl','cs','fi','fa','ku',
          'vi','th','tl','ko']
fileNames2=['110) eu_ger-w) Dutch',
'121) eu_sla-w) Czech','140)eu_ur) Finnish',
'500) as_ir) Persian','500)as_ir) Kurdish',
'530) as_2) Vietnamese','530)as_2) Thai',
'540) as_oc) Filipino','550)as_3) Korean']

langues3=['no','sv','da','uk','am','az','mg','ur',
'bn','pa','mr','gu','si','te','ml','ta','kn',
'my','km','ms','jw']
fileNames3=['111) eu_ger-n) Norwegian',
'111) eu_ger-n) Swedish','111) eu_ger-n) Danish',
'120) eu_sla-e) Ukrainian','200) sem) Amharic',
'300) tur) Azerbaijani','540) as_oc) Magalasy',
'510) as_in-n) Urdu','510) as_in-n) Bengali',
'510) as_in-n) Punjabi','510) as_in-n) Marathi',
'510) as_in-n) Gujarathi','510) as_in-n) Sinhalese',
'511) as_in-s) Telugu','511) as_in-s) Malayalam',
'511) as_in-s) Tamil','511) as_in-s) Kannada',
'520) as_1) Myanmar','530) as_2) Khmer',
'540) as_oc) Malay','540) as_oc) Javanese']

languesAll=[langues1,langues2,langues3]
fileNamesAll=[fileNames1,fileNames2,fileNames3]

##########################

def is_number(s):
    try:
        float(s)
        return True
    except:
        return False

decap = lambda s: s[:1].lower() + s[1:] if s else ''



###############################"


def loadSheet1(defaultSource):
    wb1=openpyxl.load_workbook(defaultSource)
    sheet1=wb1.worksheets[0]
    return sheet1
    
def loadAll(languesAll,c,ii,fileNamesAll,defaultSource,fileSuffix):
    langue=languesAll[ii][c]
    print '\n\n', c, langue,'\n'
    try:
        wb2=openpyxl.load_workbook(fileNamesAll[ii][c]+fileSuffix+'.xlsx')
        pb=False
    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print message
        os.chdir('..')
        wb2=openpyxl.load_workbook(defaultSource) #wb2 = deepcopy(wb1)
        os.chdir(str(ii+1))
        pb=True
    sheet2=wb2.worksheets[0]
    sheet3=wb2.worksheets[1]
    return [langue,wb2,sheet2,sheet3,pb]

def gereException(ex):
    template = "An exception of type {0} occurred. Arguments:\n{1!r}"
    message = template.format(type(ex).__name__, ex.args)
    print message
    sys.stdout.flush()        
    
def copyStyle(new_cell,old_cell):
    new_cell.font = old_cell.font
    new_cell.border = old_cell.border
    new_cell.fill = old_cell.fill
    new_cell.number_format = old_cell.number_format
    new_cell.protection = old_cell.protection
    new_cell.alignment = old_cell.alignment

def updateStyle(a,b,sheet2,sheet3,sheet1):
    print a,b
    sys.stdout.flush()
    cell1=sheet1.cell(row=b,column=a)
    cell2=sheet2.cell(row=b,column=a)
    cell3=sheet3.cell(row=b,column=a)
    for new_cell in [cell2,cell3]:
        copyStyle(new_cell,old_cell=cell1)
    return [sheet2,sheet3]


def updateCase(a,b,sheet2,sheet3,sheet1,langue):
    try:
        print a,b
        
        cell1=sheet1.cell(row=b,column=a)
        cell2=sheet2.cell(row=b,column=a)
        cell3=sheet3.cell(row=b,column=a)
        
        for new_cell in [cell2,cell3]:
            copyStyle(new_cell,old_cell=cell1)
        
        mot=cell1.value
        if (mot is not None): #and 
            if isinstance(mot, basestring):#not is_number(mot):
                bbb=translator.translate(mot, src='en', dest=langue)
                cell2.value=decap(bbb.text)
                cell3.value=decap(bbb.pronunciation)
            else:
                cell2.value=mot
                cell3.value=mot
        else:
            sheet2.cell(row=b,column=a).value=''
            sheet3.cell(row=b,column=a).value=''
        return [sheet2,sheet3]
    except Exception as ex:
        gereException(ex)
        time.sleep(1)
        [sheet2,sheet3]=updateCase(a,b,sheet2,sheet3,sheet1,langue) # retry
        return [sheet2,sheet3]


def updateRectangle(a0,a1,b0,b1,depart,defaultSource,fileSuffix,c_first=0,c_last_option=0):
    [a0_,a1_,b0_,b1_]=[a0,a1,b0,b1]
    os.chdir(depart)
    sheet1=loadSheet1(defaultSource)
    for ii in [0,1,2]:
        os.chdir(str(ii+1))
        if c_last_option:
            c_last=1
        else:
            c_last=len(languesAll[ii])
        c=c_first
        while c<c_last:
            [langue,wb2,sheet2,sheet3,pb]=loadAll(languesAll,c,ii,fileNamesAll,defaultSource,fileSuffix)
            if pb:
                [a0,a1,b0,b1]=[a0_0,a1_0,b0_0,b1_0]
            else:
                [a0,a1,b0,b1]=[a0_,a1_,b0_,b1_]
            for a in range(a0,a1+1):
                for b in range(b0,b1+1):
                    [sheet2,sheet3]=updateCase(a,b,sheet2,sheet3,sheet1,langue)
            wb2.save(fileNamesAll[ii][c]+fileSuffix+'.xlsx')#'_phrases.xlsx')
            c+=1
        os.chdir('..')


def getCoord(s):
    xy = coordinate_from_string(s) # returns ('A',4)
    col = column_index_from_string(xy[0]) # returns 1
    row = xy[1]
    return [col,row]
    
    
def updateListe(l,depart,defaultSource,fileSuffix,c_first=0,c_last_option=0):
    os.chdir(depart)
    sheet1=loadSheet1(defaultSource)
    for ii in [0,1,2]:
        os.chdir(str(ii+1))
        if c_last_option:
            c_last=1
        else:
            c_last=len(languesAll[ii])
        c=c_first
        while c<c_last:
            [langue,wb2,sheet2,sheet3,pb]=loadAll(languesAll,c,ii,fileNamesAll,defaultSource,fileSuffix)
            t=-1
            while t<len(l)-1:
                t+=1
                [a,b]=getCoord(l[t])
                [sheet2,sheet3]=updateCase(a,b,sheet2,sheet3,sheet1,langue)
            wb2.save(fileNamesAll[ii][c]+fileSuffix+'.xlsx')#'_phrases.xlsx')
            c+=1
        os.chdir('..')


def updateRectangleStyle(a0,a1,b0,b1,depart,defaultSource,fileSuffix,c_first=0,c_last_option=0):
    [a0_,a1_,b0_,b1_]=[a0,a1,b0,b1]
    os.chdir(depart)
    sheet1=loadSheet1(defaultSource)
    for ii in [0,1,2]:
        os.chdir(str(ii+1))
        if c_last_option:
            c_last=1
        else:
            c_last=len(languesAll[ii])
        c=c_first
        while c<c_last:
            [langue,wb2,sheet2,sheet3,pb]=loadAll(languesAll,c,ii,fileNamesAll,defaultSource,fileSuffix)
            if pb:
                [a0,a1,b0,b1]=[a0_0,a1_0,b0_0,b1_0]
            else:
                [a0,a1,b0,b1]=[a0_,a1_,b0_,b1_]
            for a in range(a0,a1+1):
                for b in range(b0,b1+1):
                    [sheet2,sheet3]=updateStyle(a,b,sheet2,sheet3,sheet1)
            wb2.save(fileNamesAll[ii][c]+fileSuffix+'.xlsx')#'_phrases.xlsx')
            c+=1
        os.chdir('..')


## test
#depart="C:\\Users\\Nicolas\\Desktop\\interets\\0) Langues\\0) docs generaux\\5_essai2 excel"
#a0=2; a1=2; b0=2; b1=2
#updateRectangle(a0,a1,b0,b1,firstTime,depart)

###################

[a0_0,b0_0,a1_0,b1_0]=[2,2,12,95]
depart="C:\\Users\\Nicolas\\Desktop\\interets\\0) Langues\\0) docs generaux\\5_essai2 excel"
ds='0) en.xlsx'
fs=''

[a0_0,a1_0,b0_0,b1_0]=[3,7,2,179]
depart="C:\\Users\\Nicolas\\Desktop\\interets\\0) Langues\\0) docs generaux\\6_phrases"
ds='0) en_phrases.xlsx'
fs='_phrases'

updateRectangle(a0=2 ,b0=53,a1=9 ,b1=95,depart=depart)
updateRectangle(a0=10,b0=93,a1=12,b1=95,depart=depart)

l=['E52']
updateListe(l,depart,defaultSource=ds,c_first=0,c_last_option=0,fileSuffix=fs)

updateRectangle(a0=1,b0=128,a1=2,b1=129,depart=depart,
                defaultSource=ds,c_first=0,c_last_option=0,fileSuffix=fs)

###################

os.chdir("C:\\Users\\Nicolas\\Desktop\\interets\\0) Langues\\0) docs generaux\\5_essai2 excel")


#wb1 = openpyxl.load_workbook('0)tableau_en.xlsx')
wb1 = openpyxl.load_workbook('0) en.xlsx')
sheet1=wb1.worksheets[0]
row_count=sheet1.max_row
column_count=sheet1.max_column
for b in range(10,19):
    print sheet1.cell(row=b,column=4).value

sheet1['D50'].value

for sheet in wb1.worksheets:
    print sheet.title

        
#wb2 = openpyxl.load_workbook('0)tableau_en.xlsx')
#sheet2=wb2.worksheets[0]
#sheet3=wb2.worksheets[1]

#os.chdir('3')
#os.listdir('.')
##if filename.startswith("cheese_"):
#langues=langues3
#fileNames=fileNames3
#for i in range(len(langues)):
#    fileName_old=langues[i]+'.xlsx'
#    fileName_new=fileNames[i]+'.xlsx'
#    os.rename(fileName_old, fileName_new)





#c0=len(langues)
#c=0
#while c<c0:
#    langue=langues[c]
#    print c, langue
#    wb2 = openpyxl.load_workbook(langue+'.xlsx')
#    sheet2=wb2.worksheets[0]
#    sheet3=wb2.worksheets[1]
#    for b in range(39,43):
#        mot=sheet2.cell(row=b,column=6).value
#        sheet2.cell(row=b-1,column=6).value=mot
#        mot=sheet3.cell(row=b,column=6).value
#        sheet3.cell(row=b-1,column=6).value=mot
#    for b in range(44,48):
#        mot=sheet2.cell(row=b,column=6).value
#        sheet2.cell(row=b-2,column=6).value=mot
#        mot=sheet3.cell(row=b,column=6).value
#        sheet3.cell(row=b-2,column=6).value=mot
#    for b in range(46,48):
#        sheet2.cell(row=b,column=6).value=''
#        sheet3.cell(row=b,column=6).value=''
#    wb2.save(langue+'.xlsx')
#    c+=1


#bbb=translator.translate('hello', src='en', dest='iw')
#print bbb.text
#print bbb.pronunciation



################################
#bbb=translator.translate('안녕하세요.')
# <Translated src=ko dest=en text=Good evening. pronunciation=Good evening.>
#bbb=translator.translate('안녕하세요.', dest='ja')
# <Translated src=ko dest=ja text=こんにちは。 pronunciation=Kon'nichiwa.>
#bbb=translator.translate('veritas lux mea', src='la')
# <Translated src=la dest=en text=The truth is my light pronunciation=The truth is my light>
#bbb=translator.translate('hi', src='en', dest='ru')
#print bbb.text
#print bbb.pronunciation
#############################




def coordonnees(a, b):
    return 94*(a-1)+(b-1)


def coordonnees_inv(c):
    return ((c/94)+1,(c%94)+1)






c0=len(langues)
c=0
a0=2
a1=12
b0=2
b1=95
#while c<c0:
#    langue=langues[c]
#    print '\n\n', c, langue,'\n'
#    wb2 = openpyxl.load_workbook('phrases.xlsx')
langue='0) fr'
wb2 = openpyxl.load_workbook(langue+'.xlsx')
sheet2=wb2.worksheets[0]
sheet3=wb2.worksheets[1]
for a in range(a0,a1+1):
    for b in range(b0,b1+1):
        print a,b
        mot=sheet2.cell(row=b,column=a).value
        if (mot is not None): #and 
            if isinstance(mot, basestring):#not is_number(mot):
                sheet2.cell(row=b,column=a).value=decap(mot)
                sheet3.cell(row=b,column=a).value=decap(sheet3.cell(row=b,column=a).value)
        else:
            sheet2.cell(row=b,column=a).value=''
            sheet3.cell(row=b,column=a).value=''
wb2.save(langue+'.xlsx')
#c+=1




#####################


new_sheet = workbook.create_sheet(sheetName)
default_sheet = workbook['default']

for row in default_sheet.rows:
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row_idx,
                   col=cell.col_idx, value= cell.value)
        if cell.has_style:
            new_cell.font = cell.font
            new_cell.border = cell.border
            new_cell.fill = cell.fill
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection
            new_cell.alignment = cell.alignment
