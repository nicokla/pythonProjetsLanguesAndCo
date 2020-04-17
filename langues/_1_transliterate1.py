# -*- coding: UTF-8 -*-

import time
import os
#import copy
import sys
import openpyxl
#from openpyxl.utils import coordinate_from_string, column_index_from_string
from unidecode import unidecode

os.chdir( u'/home/nicolas/Bureau/forWindows')
execfile('0) transliterate2.py')
execfile('transliterate.py')
execfile('3) transliterate_sub4_hindi.py')
execfile('4) transliterate_sub7_chinese.py')
execfile('5) transliterate_sub3_japanese5.py')
execfile('4) transliterate_sub8_kor.py')

decap = lambda s: s[:1].lower() + s[1:] if s else ''


def loadAll(fileName):
    wb=openpyxl.load_workbook(fileName+'.xlsx')
    sheet1=wb.worksheets[0]
    sheet3=wb.worksheets[2]
    return [wb,sheet1,sheet3]

def gereException(ex):
    template = "An exception of type {0} occurred. Arguments:\n{1!r}"
    message = template.format(type(ex).__name__, ex.args)
    print message
    sys.stdout.flush()        
    
def copyStyle(new_cell,old_cell):
    new_cell.font = old_cell.font
    new_cell.border = old_cell.border
    new_cell.fill = old_cell.fill
    new_cell.number_format = old_cell.number_format
    new_cell.protection = old_cell.protection
    new_cell.alignment = old_cell.alignment

#####################



def transliterate(mot,langue):
    if langue=='ru':
        return transliterateRussian(mot)
    elif langue=='ar':
        return transliterateArabic(mot)
    elif langue=='he':
        return transliterateHebrew(mot)
    elif langue=='hi':
        return transliterateHindi(mot)
    elif langue=='zh':
        return pinyin.get(mot, delimiter=" ")
    elif langue=='ja':
        return kanji_to_romaji(mot)
    elif langue=='el':
        return transliteratePolyglot(mot)
    elif langue=='th':
        return transliteratePolyglot(mot)
    elif langue=='yi':
        return transliteratePolyglot(mot)
    elif langue=='ko':
        return translit.romanize(mot)
    else:
        return mot



#######################


def updateCase(a,b,sheet1,sheet3,langue):
    try:
        print a,b
        cell1=sheet1.cell(row=b,column=a)
        cell3=sheet3.cell(row=b,column=a)
        copyStyle(new_cell=cell3,old_cell=cell1)
        mot=cell1.value
        #print mot
        if (mot is not None): #and 
            if isinstance(mot, basestring):#not is_number(mot):
                mot2=transliterate(mot,langue)
                #print mot2
                cell3.value=decap(mot2)
            else:
                cell3.value=mot
        else:
            cell3.value=''
        return sheet3
    except Exception as ex:
        gereException(ex)
        #time.sleep(1)
        #sheet3=updateCase(a,b,sheet1,sheet3,langue) # retry
        return sheet3


def transliterateRectangle(a0,b0,a1,b1,fileName,langue):
    [wb,sheet1,sheet3]=loadAll(fileName)
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            sheet3=updateCase(a,b,sheet1,sheet3,langue)
    wb.save(fileName+'.xlsx')
############################
os.chdir( u'/home/nicolas/Bureau/forWindows')

langue='ko'

[a0,b0,a1,b1]=[2,2,12,95]
fileName='550)as_3) Korean'
transliterateRectangle(a0,b0,a1,b1,fileName,langue)

[a0,b0,a1,b1]=[3,2,7,179]
fileName='550)as_3) Korean_phrases'
transliterateRectangle(a0,b0,a1,b1,fileName,langue)


                

