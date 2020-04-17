# -*- coding: UTF-8 -*-

#import time
import os
#import copy
import sys
from openpyxl import *
from openpyxl.utils import *
#from unidecode import unidecode

#cwd=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\2.0.1) Phrases, excels'
#os.chdir(cwd)

fileNames={
'en':'0) en_phrases.xlsx',
'fr':'0) fr_phrases.xlsx',
'hi':'0) hi_phrases.xlsx',
'iw':'0) iw_phrases.xlsx',
'pal':'0) pal_phrases.xlsx',
'it':'100) eu_lat) Italian_phrases.xlsx',
'pt':'100) eu_lat) Portuguese_phrases.xlsx',
'ro':'100) eu_lat) Romanian_phrases.xlsx',
'es':'100) eu_lat) Spanish_phrases.xlsx',
'nl':'110) eu_ger-w) Dutch_phrases.xlsx',
'de':'110) eu_ger-w) German_phrases.xlsx',
'yi':'110) eu_ger-w) Yiddish_phrases.xlsx',
'da':'111) eu_ger-n) Danish_phrases.xlsx',
'no':'111) eu_ger-n) Norwegian_phrases.xlsx',
'sv':'111) eu_ger-n) Swedish_phrases.xlsx',
'ru':'120) eu_sla-e) Russian_phrases.xlsx',
'uk':'120) eu_sla-e) Ukrainian_phrases.xlsx',
'cs':'121) eu_sla-w) Czech_phrases.xlsx',
'fi':'140) eu_ur) Finnish_phrases.xlsx',
'el':'160) eu_uncl) Greek_phrases.xlsx',
'am':'200) sem) Amharic_phrases.xlsx',
'ar':'200) sem) Arabic_phrases.xlsx',
'az':'300) tur) Azerbaijani_phrases.xlsx',
'tr':'300) tur) Turkish_phrases.xlsx',
'sw':'400) af) Swahili_phrases.xlsx',
'fa':'500) as_ir) Persian_phrases.xlsx',
'ku':'500) as_ir) Kurdish_phrases.xlsx',
'bn':'510) as_in-n) Bengali_phrases.xlsx',
'gu':'510) as_in-n) Gujarathi_phrases.xlsx',
'mr':'510) as_in-n) Marathi_phrases.xlsx',
'pa':'510) as_in-n) Punjabi_phrases.xlsx',
'si':'510) as_in-n) Sinhalese_phrases.xlsx',
'ur':'510) as_in-n) Urdu_phrases.xlsx',
'kn':'511) as_in-s) Kannada_phrases.xlsx',
'ml':'511) as_in-s) Malayalam_phrases.xlsx',
'ta':'511) as_in-s) Tamil_phrases.xlsx',
'te':'511) as_in-s) Telugu_phrases.xlsx',
'zh':'520) as_1) Chinese_phrases.xlsx',
'my':'520) as_1) Myanmar_phrases.xlsx',
'km':'530) as_2) Khmer_phrases.xlsx',
'vi':'530) as_2) Vietnamese_phrases.xlsx',
'th':'530) as_2) Thai_phrases.xlsx',
'tl':'540) as_oc) Filipino_phrases.xlsx',
'jw':'540) as_oc) Javanese_phrases.xlsx',
'mg':'540) as_oc) Malagasy_phrases.xlsx',
'ms':'540) as_oc) Malay_phrases.xlsx',
'ja':'550) as_3) Japanese_phrases.xlsx',
'ko':'550) as_3) Korean_phrases.xlsx'
}


decap = lambda s: s[:1].lower() + s[1:] if s else ''

langue_kn='fr'
langue_un='en'


def copyDimensions(s_to,s_from,a0,b0,a1,b1):
    for a in range(a0,a1+1):
        aa=get_column_letter(a)
        s_to.column_dimensions[aa].width=s_from.column_dimensions[aa].width
    for b in range(b0,b1+1):
        s_to.row_dimensions[b].height=s_from.row_dimensions[b].height
    s_to.page_margins.left=s_from.page_margins.left
    s_to.page_margins.right=s_from.page_margins.right
    s_to.page_margins.top=s_from.page_margins.top
    s_to.page_margins.bottom=s_from.page_margins.bottom


def copyStyleCell(new_cell,old_cell):
    new_cell.font = old_cell.font
    new_cell.border = old_cell.border
    new_cell.fill = old_cell.fill
    new_cell.number_format = old_cell.number_format
    new_cell.protection = old_cell.protection
    new_cell.alignment = old_cell.alignment

def copySheet(s_to,s_from,a0,b0,a1,b1):
    copyDimensions(s_to,s_from,a0,b0,a1,b1)
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            c_to=s_to.cell(row=b,column=a)
            c_from=s_from.cell(row=b,column=a)
            c_to.value=c_from.value
            copyStyleCell(c_to,c_from)


def tryCreateSheet(wb, sheetName):
    try:
        return wb[sheetName]
    except: #Exception as ex:
        wb.create_sheet(sheetName)
        return wb[sheetName]

listeTitres=[
'1.1) TO BE, TO HAVE',
'1.2) AUXILIARY',
'2.1) BASIC ACTIONS',
'2.2) PERCEPTION',
'2.3) MOVEMENT',
'3.1) RATIONNAL THOUGHTS',
'3.2) EMOTIONAL THOUGHTS',
'4.1) OBJECT MANIPULATION',
'4.2) EXCHANGE OF OBJECTS',
'5.1) SPEAKING',
'5.2) INTERACTING',
'6) CAUSALITY'
]

listeTitresLignes=[1,4,17,38,48,69,81,97,
123,133,157,174
]

def createDocBilingue(langue_kn,langue_un):
    wb_kn=load_workbook(fileNames[langue_kn])
    wb_un=load_workbook(fileNames[langue_un])
    wb_style=load_workbook('0) style_phrases.xlsx')
    
    s_kn=wb_kn.worksheets[0]
    s_un=wb_un.worksheets[0]
    s_n1=wb_style.worksheets[0]
    s_n2=wb_style.worksheets[1]
    
    newSheetName1='from_'+langue_kn+'_1'
    newSheetName2='from_'+langue_kn+'_2'
    s_a1=tryCreateSheet(wb_un,newSheetName1)
    s_a2=tryCreateSheet(wb_un,newSheetName2)
    
    copySheet(s_a1,s_n1,1,1,9,180)
    copySheet(s_a2,s_n2,1,1,3,180)
    
    for b in range (1,181):
        a2=7
        mot_kn2=s_kn.cell(row=b,column=a2).value
        mot_un2=s_un.cell(row=b,column=a2).value
        s_a2.cell(row=b,column=1).value=mot_kn2
        s_a2.cell(row=b,column=2).value=mot_un2
        for a in range(1,5):
            a1=a+2
            an=2*a-1
            mot_kn1=s_kn.cell(row=b,column=a1).value
            mot_un1=s_un.cell(row=b,column=a1).value
            s_a1.cell(row=b,column=an).value=mot_kn1
            s_a1.cell(row=b,column=an+1).value=mot_un1
    
    for a in range(len(listeTitresLignes)):
        i=listeTitresLignes[a]
        s_a1.cell(row=i,column=1).value=\
           listeTitres[a]
        s_a2.cell(row=i,column=1).value=\
           listeTitres[a]

    wb_un.save(fileNames[langue_un])
    print fileNames[langue_un]+' done\n'
    sys.stdout.flush()        


# Pour moi
#for s in ['ru','sw','ja','zh','de','es','it','pt',\
#'tr','el','vi','th','ko','ku','de','yi','ro','fi',\
#'cs','tl','nl','mg']:
#    createDocBilingue('en',s)

#fileNames['tl']

# with work to do (transcrire):
# ukrainien (?)   uk
# perse (?)   fa
# sinhalese  si

# with work absorber, transcrire
# polonais     pl
# hongrois (?)     hu

#createDocBilingue('fr','iw')# Pour les olims francophones
#createDocBilingue('en','iw')# Pour les olims anglosaxons
#
#
## Pour les israeliens
#createDocBilingue('pal','iw')# Pour les palestiniens
#createDocBilingue('iw','pal')
#
#
## Pour Marianne et Ankit
#createDocBilingue('en','fr')
#createDocBilingue('fr','hi')
#createDocBilingue('en','hi')




##############################


def loadAll(fileName):
    wb=openpyxl.load_workbook(fileName+'.xlsx')
    sheet1=wb.worksheets[0]
    sheet3=wb.worksheets[2]
    return [wb,sheet1,sheet3]

def gereException(ex):
    template = "An exception of type {0} occurred. Arguments:\n{1!r}"
    message = template.format(type(ex).__name__, ex.args)
    print message
    sys.stdout.flush()        
    

#####################



def transliterate(mot,langue):
    if langue=='ru':
        return transliterateRussian(mot)
    elif langue=='ar':
        return transliterateArabic(mot)
    elif langue=='he':
        return transliterateHebrew(mot)
    elif langue=='hi':
        return transliterateHindi(mot)
    elif langue=='zh':
        return pinyin.get(mot, delimiter=" ")
    elif langue=='ja':
        return kanji_to_romaji(mot)
    elif langue=='el':
        return transliteratePolyglot(mot)
    elif langue=='th':
        return transliteratePolyglot(mot)
    elif langue=='yi':
        return transliteratePolyglot(mot)
    elif langue=='ko':
        return translit.romanize(mot)
    else:
        return mot



#######################


def updateCase(a,b,sheet1,sheet3,langue):
    try:
        print a,b
        cell1=sheet1.cell(row=b,column=a)
        cell3=sheet3.cell(row=b,column=a)
        copyStyle(new_cell=cell3,old_cell=cell1)
        mot=cell1.value
        #print mot
        if (mot is not None): #and 
            if isinstance(mot, basestring):#not is_number(mot):
                mot2=transliterate(mot,langue)
                #print mot2
                cell3.value=decap(mot2)
            else:
                cell3.value=mot
        else:
            cell3.value=''
        return sheet3
    except Exception as ex:
        gereException(ex)
        #time.sleep(1)
        #sheet3=updateCase(a,b,sheet1,sheet3,langue) # retry
        return sheet3


def transliterateRectangle(a0,b0,a1,b1,fileName,langue):
    [wb,sheet1,sheet3]=loadAll(fileName)
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            sheet3=updateCase(a,b,sheet1,sheet3,langue)
    wb.save(fileName+'.xlsx')
############################
#os.chdir( u'/home/nicolas/Bureau/forWindows')
#
#langue='ko'
#
#[a0,b0,a1,b1]=[2,2,12,95]
#fileName='550)as_3) Korean'
#transliterateRectangle(a0,b0,a1,b1,fileName,langue)
#
#[a0,b0,a1,b1]=[3,2,7,179]
#fileName='550)as_3) Korean_phrases'
#transliterateRectangle(a0,b0,a1,b1,fileName,langue)


                

