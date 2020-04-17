# -*- coding: UTF-8 -*-

#import time
import os
#import copy
import sys
from openpyxl import *
from openpyxl.utils import *
#from unidecode import unidecode
import os.path

fileNames={
'en':'0) en',
'fr':'0) fr',
'hi':'0) hi',
'iw':'0) iw',
'iwAlph':'0) iwAlph',
'pal':'0) pal',
'it':'100) eu_lat) Italian',
'pt':'100) eu_lat) Portuguese',
'ro':'100) eu_lat) Romanian',
'es':'100) eu_lat) Spanish',
'nl':'110) eu_ger-w) Dutch',
'de':'110) eu_ger-w) German',
'yi':'110) eu_ger-w) Yiddish',
'da':'111) eu_ger-n) Danish',
'no':'111) eu_ger-n) Norwegian',
'sv':'111) eu_ger-n) Swedish',
'ru':'120) eu_sla-e) Russian',
'uk':'120) eu_sla-e) Ukrainian',
'cs':'121) eu_sla-w) Czech',
'fi':'140) eu_ur) Finnish',
'el':'160) eu_uncl) Greek',
'am':'200) sem) Amharic',
'ar':'200) sem) Arabic',
'az':'300) tur) Azerbaijani',
'tr':'300) tur) Turkish',
'sw':'400) af) Swahili',
'fa':'500) as_ir) Persian',
'ku':'500) as_ir) Kurdish',
'bn':'510) as_in-n) Bengali',
'gu':'510) as_in-n) Gujarathi',
'mr':'510) as_in-n) Marathi',
'pa':'510) as_in-n) Punjabi',
'si':'510) as_in-n) Sinhalese',
'ur':'510) as_in-n) Urdu',
'kn':'511) as_in-s) Kannada',
'ml':'511) as_in-s) Malayalam',
'ta':'511) as_in-s) Tamil',
'te':'511) as_in-s) Telugu',
'zh':'520) as_1) Chinese',
'my':'520) as_1) Myanmar',
'km':'530) as_2) Khmer',
'vi':'530) as_2) Vietnamese',
'th':'530) as_2) Thai',
'tl':'540) as_oc) Filipino',
'jw':'540) as_oc) Javanese',
'mg':'540) as_oc) Malagasy',
'ms':'540) as_oc) Malay',
'ja':'550) as_3) Japanese',
'ko':'550) as_3) Korean'
}


decap = lambda s: s[:1].lower() + s[1:] if s else ''

langue_kn='fr'
langue_un='pal'


def copyDimensions(s_to,s_from,a0,b0,a1,b1):
    for a in range(a0,a1+1):
        aa=get_column_letter(a)
        s_to.column_dimensions[aa].width=s_from.column_dimensions[aa].width
    for b in range(b0,b1+1):
        s_to.row_dimensions[b].height=s_from.row_dimensions[b].height
    s_to.page_margins.left=s_from.page_margins.left
    s_to.page_margins.right=s_from.page_margins.right
    s_to.page_margins.top=s_from.page_margins.top
    s_to.page_margins.bottom=s_from.page_margins.bottom


def copyStyleCell(new_cell,old_cell):
    new_cell.font = old_cell.font
    new_cell.border = old_cell.border
    new_cell.fill = old_cell.fill
    new_cell.number_format = old_cell.number_format
    new_cell.protection = old_cell.protection
    new_cell.alignment = old_cell.alignment

def copySheet(s_to,s_from,a0,b0,a1,b1):
    copyDimensions(s_to,s_from,a0,b0,a1,b1)
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            c_to=s_to.cell(row=b,column=a)
            c_from=s_from.cell(row=b,column=a)
            c_to.value=c_from.value
            copyStyleCell(c_to,c_from)


def tryCreateSheet(wb, sheetName):
    try:
        return wb[sheetName]
    except: #Exception as ex:
        wb.create_sheet(sheetName)
        return wb[sheetName]

def createDocBilingue(langue_kn,langue_un,outputFileName):
    cwd_mots=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\1.0.1) Mots, excels'
    os.chdir(cwd_mots)

    wb_kn=load_workbook(fileNames[langue_kn]+'.xlsx')
    wb_un=load_workbook(fileNames[langue_un]+'.xlsx')
    wb_style=load_workbook('0) style bilingue.xlsx')
    
    s_kn=wb_kn.worksheets[0]
    s_un=wb_un.worksheets[0]
    s_n1=wb_style.worksheets[0]
    s_n2=wb_style.worksheets[1]
    
    newSheetName1='mots_1'
    newSheetName2='mots_2'
    wb_output=Workbook() #load_workbook(outputFileName+'.xlsx')
    s_a1=tryCreateSheet(wb_output,newSheetName1)
    s_a2=tryCreateSheet(wb_output,newSheetName2)
    
    copySheet(s_a1,s_n1,1,1,11,94)
    copySheet(s_a2,s_n2,1,1,11,94)
    
    for a in range(1,6):
        for b in range (1,95):
            a1=a+1
            b1=b+1
            an=2*a-1
            bn=b
            mot_kn1=s_kn.cell(row=b1,column=a1).value
            mot_un1=s_un.cell(row=b1,column=a1).value
            s_a1.cell(row=bn,column=an).value=mot_kn1
            s_a1.cell(row=bn,column=an+1).value=mot_un1
    
            a2=a+7
            mot_kn2=s_kn.cell(row=b1,column=a2).value
            mot_un2=s_un.cell(row=b1,column=a2).value
            s_a2.cell(row=bn,column=an).value=mot_kn2
            s_a2.cell(row=bn,column=an+1).value=mot_un2
    
    #######################
    # on passe aux phrases
    cwd_phrases=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\2.0.1) Phrases, excels'
    os.chdir(cwd_phrases)

    wb_kn=load_workbook(fileNames[langue_kn]+'_phrases'+'.xlsx')
    wb_un=load_workbook(fileNames[langue_un]+'_phrases'+'.xlsx')
    wb_style=load_workbook('0) style phrases.xlsx')

    s_kn=wb_kn.worksheets[0]
    s_un=wb_un.worksheets[0]
    s_n1=wb_style.worksheets[0]
    s_n2=wb_style.worksheets[1]

    newSheetName1='phrases_1'
    newSheetName2='phrases_2'
    s_a1=tryCreateSheet(wb_output,newSheetName1)
    s_a2=tryCreateSheet(wb_output,newSheetName2)
    
    copySheet(s_a1,s_n1,1,1,9,180)
    copySheet(s_a2,s_n2,1,1,4,180)
    
    for b in range (1,181):
        a2=7
        mot_kn2=s_kn.cell(row=b,column=a2).value
        mot_un2=s_un.cell(row=b,column=a2).value
        s_a2.cell(row=b,column=2).value=mot_kn2
        s_a2.cell(row=b,column=3).value=mot_un2
        for a in range(1,5):
            a1=a+2
            an=2*a-1
            mot_kn1=s_kn.cell(row=b,column=a1).value
            mot_un1=s_un.cell(row=b,column=a1).value
            s_a1.cell(row=b,column=an+1).value=mot_kn1
            s_a1.cell(row=b,column=an+2).value=mot_un1
    
    # save file
    cwd_output=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand'
    os.chdir(cwd_output)
    wb_output.save(outputFileName+'.xlsx')
    print langue_kn +' to '+ langue_un +' : done'
    sys.stdout.flush()        


def createDocBilingue_noErase(langue_kn,langue_un,outputFileName):
    cwd_output=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand'
    os.chdir(cwd_output)
    if(not os.path.isfile(outputFileName+'.xlsx')):
        createDocBilingue(langue_kn,langue_un,outputFileName)
    else:
        print langue_kn +' to '+ langue_un +' : done already'
        sys.stdout.flush()        


# Pour moi
#for s in ['pal','ru','sw','ja''zh','de','es','it','pt',\
#'tr','el','vi','th','ko','ku','de','yi','ro','fi',\
#'cs','tl','nl','mg']:
#    createDocBilingue('en',s)
#
#fileNames['tl']

#createDocBilingue('fr','en','exemple')

for s1 in ['en','fr','iw','iwAlph']:
    for s in fileNames:
        if(s!=s1):
            createDocBilingue_noErase(s1,s,s1+'_to_'+s)

## Pour les olims francophones
#createDocBilingue('fr','iw')
#
## Pour les olims anglosaxons
#createDocBilingue('en','iw')
#
## Pour les israeliens
#createDocBilingue('iw','pal')
#
## Pour les palestiniens
#createDocBilingue('pal','iw')
#
## Pour Marianne et Ankit
#createDocBilingue('en','fr')
#createDocBilingue('fr','hi')
#createDocBilingue('en','hi')





# with work to do (transcrire):
# ukrainien (?)   uk
# perse (?)   fa
# sinhalese  si

# with work absorber, transcrire
# polonais     pl
# hongrois (?)     hu


################################################
######################################################

import win32com.client
import glob

cwdRoot=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue'
cwdFiles=cwdRoot+r'\3) all grand'

o = win32com.client.Dispatch("Excel.Application")
o.Visible = False

def createPdf(nameXlsx, listeNameTabs, namePdf):
    wb = o.Workbooks.Open(nameXlsx)
    l=[]
    count = wb.Sheets.Count
    for i in range(1,count+1):
        ws = wb.Sheets(i)
        if (ws.Name in listeNameTabs):
            l.append(i)
    
    wb.WorkSheets(l).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, namePdf)
    o.Workbooks.Close()

def remetVisible():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True
    o.Quit()
    del o


def xlsxToPdfInDirectory(directory):
    l=(glob.glob(cwdFiles+r'\*.xlsx'))
    listeNameTabs=['mots_1','mots_2','phrases_1','phrases_2']
    for f in l:
        if(not os.path.isfile(f[:-5]+'.pdf')):
            print f[:-5]+' : start'
            sys.stdout.flush()
            createPdf(f, listeNameTabs, f[:-5]+'.pdf')
            print f[:-5]+' : done\n'
            sys.stdout.flush() 
    remetVisible()

xlsxToPdfInDirectory(cwdFiles)



