# -*- coding: UTF-8 -*-

#import time
import os
#import copy
import sys
from openpyxl import *
from openpyxl.utils import *
#from unidecode import unidecode
import os.path

fileNames={
'en':'0) en',
'fr':'0) fr',
'hi':'0) hi',
'iw':'0) iw',
'iwAlph':'0) iwAlph',
'pal':'0) pal',
'it':'100) eu_lat) Italian',
'pt':'100) eu_lat) Portuguese',
'ro':'100) eu_lat) Romanian',
'es':'100) eu_lat) Spanish',
'nl':'110) eu_ger-w) Dutch',
'de':'110) eu_ger-w) German',
'yi':'110) eu_ger-w) Yiddish',
'da':'111) eu_ger-n) Danish',
'no':'111) eu_ger-n) Norwegian',
'sv':'111) eu_ger-n) Swedish',
'ru':'120) eu_sla-e) Russian',
'uk':'120) eu_sla-e) Ukrainian',
'cs':'121) eu_sla-w) Czech',
'fi':'140) eu_ur) Finnish',
'el':'160) eu_uncl) Greek',
'am':'200) sem) Amharic',
'ar':'200) sem) Arabic',
'az':'300) tur) Azerbaijani',
'tr':'300) tur) Turkish',
'sw':'400) af) Swahili',
'fa':'500) as_ir) Persian',
'ku':'500) as_ir) Kurdish',
'bn':'510) as_in-n) Bengali',
'gu':'510) as_in-n) Gujarathi',
'mr':'510) as_in-n) Marathi',
'pa':'510) as_in-n) Punjabi',
'si':'510) as_in-n) Sinhalese',
'ur':'510) as_in-n) Urdu',
'kn':'511) as_in-s) Kannada',
'ml':'511) as_in-s) Malayalam',
'ta':'511) as_in-s) Tamil',
'te':'511) as_in-s) Telugu',
'zh':'520) as_1) Chinese',
'my':'520) as_1) Myanmar',
'km':'530) as_2) Khmer',
'vi':'530) as_2) Vietnamese',
'th':'530) as_2) Thai',
'tl':'540) as_oc) Filipino',
'jw':'540) as_oc) Javanese',
'mg':'540) as_oc) Malagasy',
'ms':'540) as_oc) Malay',
'ja':'550) as_3) Japanese',
'ko':'550) as_3) Korean'
}


decap = lambda s: s[:1].lower() + s[1:] if s else ''

langue_kn='fr'
langue_un='iw'


cwd2=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand\sub'
cwd1=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand'
os.chdir(cwd1)
wb0=load_workbook("coucou.xlsx")
s0=wb0.worksheets[0]
cell0=s0.cell(row=1,column=1)
#cell0.font.size
l=(glob.glob(cwd1+r'\*.xlsx'))
for f in l:
    fileName=f[:-5]
    wb1=load_workbook(fileName+".xlsx")
    n=len(wb1.worksheets)
    #a=1
    for a in range(n):
        s=wb1.worksheets[a]
        mc=s.max_column
        mr=s.max_row
        s.row_dimensions[1].height
        #b=1
        for b in range(1,mr+1):
            s.row_dimensions[b].height=9 #18
            #c=1
            for c in range(1,mc+1):
                cell=s.cell(row=b,column=c)
                #cell.font.size=7 #11
                cell.font=cell0.font
    os.chdir(cwd2)
    x=len(fileName)
    while fileName[x-1]!='\\':
        x-=1
    wb1.save(fileName[x:]+'_short.xlsx')
    os.chdir(cwd1)


################################################
######################################################

import win32com.client
import glob

o = win32com.client.Dispatch("Excel.Application")
o.Visible = False

def createPdf(nameXlsx, listeNameTabs, namePdf):
    wb = o.Workbooks.Open(nameXlsx)
    l=[]
    count = wb.Sheets.Count
    for i in range(1,count+1):
        ws = wb.Sheets(i)
        if (ws.Name in listeNameTabs):
            l.append(i)
    
    wb.WorkSheets(l).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, namePdf)
    o.Workbooks.Close()

def remetVisible():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True
    o.Quit()
    del o


def xlsxToPdfInDirectory(directory):
    l=(glob.glob(directory+r'\*.xlsx'))
    listeNameTabs=['mots_1','mots_2','phrases_1','phrases_2']
    for f in l:
        if(not os.path.isfile(f[:-5]+'.pdf')):
            print f[:-5]+' : start'
            sys.stdout.flush()
            createPdf(f, listeNameTabs, f[:-5]+'.pdf')
            print f[:-5]+' : done\n'
            sys.stdout.flush() 
    remetVisible()

xlsxToPdfInDirectory(cwd2)



