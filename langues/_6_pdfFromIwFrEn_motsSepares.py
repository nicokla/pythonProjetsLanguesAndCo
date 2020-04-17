# -*- coding: UTF-8 -*-

#import time
import os
#import copy
import sys
from openpyxl import *
from openpyxl.utils import *
#from unidecode import unidecode
import os.path
import glob


fileNames={
'en':'0) en',
'fr':'0) fr',
'hi':'0) hi',
'iw':'0) iw',
'iwAlph':'0) iwAlph',
'pal':'0) pal',
'it':'100) eu_lat) Italian',
'pt':'100) eu_lat) Portuguese',
'ro':'100) eu_lat) Romanian',
'es':'100) eu_lat) Spanish',
'nl':'110) eu_ger-w) Dutch',
'de':'110) eu_ger-w) German',
'yi':'110) eu_ger-w) Yiddish',
'da':'111) eu_ger-n) Danish',
'no':'111) eu_ger-n) Norwegian',
'sv':'111) eu_ger-n) Swedish',
'ru':'120) eu_sla-e) Russian',
'uk':'120) eu_sla-e) Ukrainian',
'cs':'121) eu_sla-w) Czech',
'fi':'140) eu_ur) Finnish',
'el':'160) eu_uncl) Greek',
'am':'200) sem) Amharic',
'ar':'200) sem) Arabic',
'az':'300) tur) Azerbaijani',
'tr':'300) tur) Turkish',
'sw':'400) af) Swahili',
'fa':'500) as_ir) Persian',
'ku':'500) as_ir) Kurdish',
'bn':'510) as_in-n) Bengali',
'gu':'510) as_in-n) Gujarathi',
'mr':'510) as_in-n) Marathi',
'pa':'510) as_in-n) Punjabi',
'si':'510) as_in-n) Sinhalese',
'ur':'510) as_in-n) Urdu',
'kn':'511) as_in-s) Kannada',
'ml':'511) as_in-s) Malayalam',
'ta':'511) as_in-s) Tamil',
'te':'511) as_in-s) Telugu',
'zh':'520) as_1) Chinese',
'my':'520) as_1) Myanmar',
'km':'530) as_2) Khmer',
'vi':'530) as_2) Vietnamese',
'th':'530) as_2) Thai',
'tl':'540) as_oc) Filipino',
'jw':'540) as_oc) Javanese',
'mg':'540) as_oc) Malagasy',
'ms':'540) as_oc) Malay',
'ja':'550) as_3) Japanese',
'ko':'550) as_3) Korean'
}


decap = lambda s: s[:1].lower() + s[1:] if s else ''

langue_kn='fr'
langue_un='iw'



def copyContent(s_from,a0,b0,a1,b1,
                s_to,  a2,b2): #,a3,b3
    da=a2-a0
    db=b2-b0
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            s_to.cell(row=b+db,column=a+da).value= \
            s_from.cell(row=b,column=a).value


# s_from,a0,b0,a1,b1,s_to,a2,b2
m_old=[
[2,1,1,4,27,    0,1,3],
[2,9,18,10,27,  0,1,32],
[2,5,1,8,27,    0,1,55],
[1,1,1,4,27,    1,1,3],
[1,5,22,6,27,1,3,31],
[1,5,1,6,21,1,1,31],
[1,7,1,10,27,1,1,53],
[1,5,53,8,90,2,1,3],
[1,1,53,4,90,3,1,3],
[2,5,28,8,52,4,1,3],
[2,9,28,10,41,4,1,29],
[2,9,42,10,52,4,3,29],
[2,9,1,10,17,5,1,3],
[1,1,91,4,94,5,1,22],
[1,1,28,2,29,5,1,33],
[1,5,90,8,94,5,1,27],
[1,9,53,10,74,6,1,3],
[2,1,53,2,75,6,3,3],
[2,3,53,4,67,6,1,27],
[2,3,68,4,75,6,3,27],
[1,9,75,10,94,7,1,3],
[2,1,76,2,94,7,3,4],
[2,3,76,4,94,7,1,25],
[2,5,53,8,94,8,1,3],
[2,9,53,10,94,8,1,51]
]


m=[
[2,1,1,4,27,    0,1,3],
[2,9,18,10,27,  0,1,32],
[2,5,1,8,27,    1,1,1],
[1,1,1,4,27,    2,1,3],
[1,5,22,6,27,2,3,31],
[1,5,1,6,21,2,1,31],
[1,7,1,10,27,3,1,1],
[1,5,53,8,90,4,1,3],
[1,1,53,4,90,5,1,3],
[2,5,28,8,52,6,1,3],
[2,9,28,10,41,6,1,29],
[2,9,42,10,52,6,3,29],
[2,9,1,10,17,7,1,3],
[1,1,91,4,94,7,1,22],
[1,1,28,2,29,7,1,33],
[1,5,90,8,94,7,1,27],
[1,9,53,10,74,8,1,3],
[2,1,53,2,75,8,3,3],
[2,3,53,4,67,8,1,27],
[2,3,68,4,75,8,3,27],
[1,9,75,10,94,9,1,3],
[2,1,76,2,94,9,3,4],
[2,3,76,4,94,9,1,25],
[2,5,53,8,94,10,1,3],
[2,9,53,10,94,11,1,1]
]


cwd2=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand\sub2'
cwd1=r'C:\Users\Nicolas\Desktop\interets\0) Langues\multilingue\3) all grand'
l=(glob.glob(cwd1+r'\*.xlsx'))

f=l[3]
for f in l[3:]:
    os.chdir(cwd1)
    wb0=load_workbook("coucou.xlsx")
    w0=wb0.worksheets
    #s0=w0[0]
    #cell0=s0.cell(row=1,column=1)
    fileName=f[:-5]
    wb1=load_workbook(fileName+".xlsx")
    w1=wb1.worksheets
    #n=len(wb1.worksheets)
    a=0
    for a in range(len(m)):
        ll=m[a]
        copyContent(w1[ll[0]],ll[1],ll[2],ll[3],ll[4],\
                    w0[ll[5]],ll[6],ll[7])
    w0[4].cell(row=40,column=3).value=""
    w0[4].cell(row=40,column=4).value=""
    w0[7].cell(row=27,column=1).value=""
    w0[7].cell(row=27,column=2).value=""
    os.chdir(cwd2)
    x=len(fileName)
    while fileName[x-1]!='\\':
        x-=1
    wb0.save(fileName[x:]+'_motsSep.xlsx')
    print fileName[x:]+' : done'
    sys.stdout.flush()


################################################
######################################################

import win32com.client

o = win32com.client.Dispatch("Excel.Application")
o.Visible = False

def createPdf(nameXlsx, listeNameTabs, namePdf):
    wb = o.Workbooks.Open(nameXlsx)
    l=[]
    count = wb.Sheets.Count
    for i in range(1,count+1):
        ws = wb.Sheets(i)
        if (ws.Name in listeNameTabs):
            l.append(i)
    
    wb.WorkSheets(l).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, namePdf)
    o.Workbooks.Close()

def remetVisible():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True
    o.Quit()
    del o


def xlsxToPdfInDirectory(directory, listeNameTabs):
    l=(glob.glob(directory+r'\*.xlsx'))
    
    for f in l:
        if(not os.path.isfile(f[:-5]+'.pdf')):
            print f[:-5]+' : start'
            sys.stdout.flush()
            createPdf(f, listeNameTabs, f[:-5]+'.pdf')
            print f[:-5]+' : done\n'
            sys.stdout.flush() 
    remetVisible()
    
listeNameTabs=['ma','ma2','mi','mi2','eifo','matay','eikh','eize','kama',\
'divers','medias','medias2']
xlsxToPdfInDirectory(cwd2, listeNameTabs)



