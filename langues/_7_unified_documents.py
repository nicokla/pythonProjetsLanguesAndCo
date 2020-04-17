# unified docs


import os
import sys
from openpyxl import *
from openpyxl.utils import *
import glob
import copy

languagesNamesDoneButNotToKeep={
'iwAlph':'Hebrew (hebrew alphabet)',
'ar':'Arabic'
}

languagesNamesToDo={
'pal':'Levantine Arabic',  #todo : changer pal vers 2 lettres (aussi file names)
'uk':'Ukrainian',
'am':'Amharic',
'fa':'Persian',
'bn':'Bengali',
'gu':'Gujarathi',
'mr':'Marathi',
'pa':'Punjabi',
'si':'Sinhalese',
'ur':'Urdu',
'kn':'Kannada',
'ml':'Malayalam',
'ta':'Tamil',
'te':'Telugu',
'my':'Myanmar',
'km':'Khmer'
}

languagesNamesOK={
'en':'English',
'fr':'French',
'hi':'Hindi',
'iw':'Hebrew',
'it':'Italian',
'pt':'Portuguese',
'ro':'Romanian',
'es':'Spanish',
'nl':'Dutch',
'de':'German',
'yi':'Yiddish',
'da':'Danish',
'no':'Norwegian',
'sv':'Swedish',
'ru':'Russian',
'cs':'Czech',
'fi':'Finnish',
'el':'Greek',
'az':'Azerbaijani',
'tr':'Turkish',
'sw':'Swahili',
'ku':'Kurdish',
'zh':'Chinese',
'vi':'Vietnamese',
'th':'Thai',
'tl':'Filipino',
'jw':'Javanese',
'mg':'Malagasy',
'ms':'Malay',
'ja':'Japanese',
'ko':'Korean'
}



def copyDimensions(s_to,s_from,a0,b0,a1,b1):
    for a in range(a0,a1+1):
        aa=get_column_letter(a)
        s_to.column_dimensions[aa].width=s_from.column_dimensions[aa].width
    for b in range(b0,b1+1):
        s_to.row_dimensions[b].height=s_from.row_dimensions[b].height
    s_to.page_margins.left=s_from.page_margins.left
    s_to.page_margins.right=s_from.page_margins.right
    s_to.page_margins.top=s_from.page_margins.top
    s_to.page_margins.bottom=s_from.page_margins.bottom


def copyStyleCell(new_cell,old_cell):
    new_cell.font = old_cell.font
    new_cell.border = old_cell.border
    new_cell.fill = old_cell.fill
    new_cell.number_format = old_cell.number_format
    new_cell.protection = old_cell.protection
    new_cell.alignment = old_cell.alignment

def copyContent(s_to,s_from,l):
    [a0,b0,a1,b1]=l
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            c_to=s_to.cell(row=b,column=a)
            c_from=s_from.cell(row=b,column=a)
            c_to.value=c_from.value
            #print(str(a) + ' ' + str(b))
            #sys.stdout.flush()
            #copyStyleCell(c_to,c_from)

def copyContentAuDebut(s_to,s_from,l):
    [a0,b0,a1,b1]=l
    for a in range(a0,a1+1):
        for b in range(b0,b1+1):
            c_to=s_to.cell(row=b-b0+1,column=a-a0+1)
            c_from=s_from.cell(row=b,column=a)
            c_to.value=c_from.value

def copySheet(s_to,s_from,a0,b0,a1,b1):
    copyDimensions(s_to,s_from,a0,b0,a1,b1)
    copyContent(s_to,s_from,[a0,b0,a1,b1])

def copyDimensionsAll(s_to,s_from):
    a1 = s_from.max_column
    b1 = s_from.max_row
    copyDimensions(s_to,s_from,1,1,a1,b1)

def copyContentAll(s_to,s_from): #[s_to,s_from] = [sheet0,sheet1]
    a1 = s_from.max_column
    b1 = s_from.max_row
    copyContent(s_to,s_from,[1,1,a1,b1])


def copyColumn(s_to,s_from,colList):
    b1 = s_from.raw_max
    for a in colList:
        for b in range(1,b1+1):
            c_to=s_to.cell(row=b,column=a)
            c_from=s_from.cell(row=b,column=a)
            c_to.value=c_from.value
            copyStyleCell(c_to,c_from)



path1 = '/Users/nicolas/Desktop/langues/multilingue/3) all grand/mots/'
path2 = '/Users/nicolas/Desktop/langues/multilingue/3) all grand/phrases/'
path3 = '/Users/nicolas/Desktop/langues/multilingue/3) all grand/bon ordre final/'
pathRoot = '/Users/nicolas/Desktop/langues/multilingue/3) all grand/'

os.chdir(path1)
for file in glob.glob("*.xlsx"):
    print(file)
os.chdir(path2)
workbook000 = load_workbook(pathRoot+'en_to_fr_motsSep.xlsx')

# 1-47
# 48-95
# 97-141
# 142-180


l=[3,4,3,4,3,4,3,4]
l2=[[1,1,9,47],
    [1,1,3,47],
    [1,48,9,95],
    [1,48,3,95],
    [1,97,9,141],
    [1,97,3,141],
    [1,142,9,180],
    [1,142,3,180]
]

for aaa in ['en', 'fr', 'iw']:    #aaa='en'   
    for bbb in languagesNamesOK:     #bbb = 'iw'
        if bbb != aaa:
            workbook0 = load_workbook(pathRoot+'en_to_fr_motsSep.xlsx') #copy.deepcopy(workbook000)
            os.chdir(path1)
            workbook1 = load_workbook(aaa + '_to_' + bbb + '_motsSep.xlsx')
            os.chdir(path2)
            workbook2 = load_workbook(aaa + '_to_' + bbb + '.xlsx')
            
            sheet0 = workbook0.worksheets[0]
            sheet0.cell(row=6,column=3).value = languagesNamesOK[aaa]
            sheet0.cell(row=8,column=3).value = languagesNamesOK[bbb]
            for i in range(1,13): #i=1
                sheet0 = workbook0.worksheets[i]
                sheet1 = workbook1.worksheets[i-1]
                copyContentAll(sheet0,sheet1)
            for i in range(13,13+8):
                sheet0 = workbook0.worksheets[i]
                sheet2 = workbook2.worksheets[l[i-13]]
                copyContentAuDebut(sheet0,sheet2,l2[i-13])
                
            os.chdir(path3)
            name3 = (languagesNamesOK[aaa] + ' to ' + languagesNamesOK[bbb] + '.xlsx')
            workbook0.save(name3)
            print(name3)
            sys.stdout.flush()










