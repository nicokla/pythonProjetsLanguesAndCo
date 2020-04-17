import os
import sys
from openpyxl import *
from openpyxl.utils import *
from openpyxl.styles import Alignment
import os.path
import glob
import win32com.client


decap = lambda s: s[:1].lower() + s[1:] if s else ''

def isHeb(s):
    for c in s:
        if ord(c) >= 1424 and ord(c) <= 1514:
            return True
    return False

cwd1=r'/Users/nicolas/Desktop/langues/multilingue/3) all grand'
l1=(glob.glob(cwd1+r'/iwAlph*.xlsx'))
cwd2=r'/Users/nicolas/Desktop/langues/multilingue/3) all grand/sub2'
l2=(glob.glob(cwd2+r'/iwAlph*.xlsx'))

os.chdir(cwd2)
for f in l2:
    wb1=load_workbook(f)
    for s in wb1.worksheets:
        for i in range(1,s.max_row+1):
            for j in range(1,s.max_column+1):
                t=s.cell(row=i,column=j)
                if t.value != None:
                    if isHeb(t.value):
                        t.alignment = Alignment(horizontal="right")
    wb1.save(f)
    print(f)
    sys.stdout.flush()


def createPdf(nameXlsx, namePdf):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False #True
    wb = o.Workbooks.Open(nameXlsx)
    l=[]
    count = wb.Sheets.Count
    for i in range(1,count+1):
        ws = wb.Sheets(i)
        #if (ws.Name in listeNameTabs):
        l.append(i)
    wb.WorkSheets(l).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, namePdf)
    o.Quit()
    del o


def remetVisible():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True
    o.Quit()
    del o

for f in l1:
    f=l1[0]
    createPdf(f,f[:-5]+'.pdf')


################################################
######################################################

import win32com.client

o = win32com.client.Dispatch("Excel.Application")
o.Visible = False

def createPdf(nameXlsx, listeNameTabs, namePdf):
    wb = o.Workbooks.Open(nameXlsx)
    l=[]
    count = wb.Sheets.Count
    for i in range(1,count+1):
        ws = wb.Sheets(i)
        if (ws.Name in listeNameTabs):
            l.append(i)
    
    wb.WorkSheets(l).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, namePdf)
    o.Workbooks.Close()

def remetVisible():
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True
    o.Quit()
    del o


def xlsxToPdfInDirectory(directory, listeNameTabs):
    l=(glob.glob(directory+r'\*.xlsx'))
    
    for f in l:
        if(not os.path.isfile(f[:-5]+'.pdf')):
            print f[:-5]+' : start'
            sys.stdout.flush()
            createPdf(f, listeNameTabs, f[:-5]+'.pdf')
            print f[:-5]+' : done\n'
            sys.stdout.flush() 
    remetVisible()
    
listeNameTabs=['ma','ma2','mi','mi2','eifo','matay','eikh','eize','kama',\
'divers','medias','medias2']
xlsxToPdfInDirectory(cwd2, listeNameTabs)



