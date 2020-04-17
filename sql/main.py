##########
# import xls

import openpyxl as px
import numpy as np
import os

W = px.load_workbook('/home/nicolas/Bureau/job/data2.xlsx', use_iterators = True)
print W.get_sheet_names()
p = W.get_sheet_by_name(name = 'HISTORIQUE STAGES SUPAERO_MASTE')

#p['D2'].value
rowSize=p.get_highest_row()
colSize=p.get_highest_column()

a=[]
for row in p.iter_rows():
    for k in row:
        a.append(k.value)

aa= np.resize(a, [rowSize, colSize])
aa0=aa[1:,:]
aa1=aa[0,:]


#####################
# put in the sql database

from unidecode import unidecode
#print unidecode(u"\u5317\u4EB0")

import sqlite3 as lite

try: os.remove("truc.db")
except: pass
con = lite.connect("truc.db") #':memory:'
cur = con.cursor()

cur.execute("""CREATE TABLE stages(pays TEXT,
            organisme TEXT, secteur TEXT,
            typeDeStage TEXT,
            etudiant TEXT,
            promo INT,
            diplome TEXT,
            domaine TEXT,
            appro TEXT,
            sujet TEXT,
            datesStages TEXT,
            mailTuteur TEXT,
            tuteur TEXT)""")
#cur.execute("DROP TABLE stages")

toto=[0,1,2,3,4,5,6,7,8,9,11,13,14]
rr=[]
for i in range(aa0.shape[0]):
    try :
        cur.execute(unidecode('INSERT INTO stages VALUES("%s", "%s", "%s", "%s", "%s",  %d, "%s", "%s", "%s", "%s", "%s", "%s", "%s")' % tuple(aa0[i,toto])))
    except:
        print i
        rr.append(i)
        pass
float(rr.__len__())/aa0.shape[0]

def showForDebug(rr):
    for i in rr:
        print(unidecode('INSERT INTO stages VALUES("%s", "%s", "%s", "%s", "%s",  %d, "%s", "%s", "%s", "%s", "%s", "%s", "%s")' % tuple(aa0[i,toto])))


rr2=[]
for i in rr:
    k=(aa0[i,toto])
    for j in range(k.__len__()):
        if(isinstance(k[j], basestring)):
            k[j]=unidecode(k[j]);
            k[j]=k[j].replace('"', "'")
        elif(k[j] is None):
            k[j]=-1
    print k
    try :
        cur.execute(unidecode('INSERT INTO stages VALUES("%s", "%s", "%s", "%s", "%s",  %d, "%s", "%s", "%s", "%s", "%s", "%s", "%s")' % tuple(k)))
    except:
        rr2.append(i)
        pass
rr2.__len__()
float(rr2.__len__())/rr.__len__()

#showForDebug(rr2[4:])

# bb=tuple(aa0[4,toto])
# cur.execute("INSERT INTO stages VALUES('%s', '%s', '%s', '%s', '%s',  %d, '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % bb)
# print "INSERT INTO stages VALUES('%s', '%s', '%s', '%s', '%s',  %d, '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % bb
# INSERT INTO stages VALUES('FRANCE', 'AEVA ASSOCIATION ETUDIANTE VALDOIDIENNE D’ARCHEOLOGIE', 'SPORTS - TOURISME -  LOISIRS',
#  'Découverte de l'entreprise', 'Adrien   DAGALLIER',  2014, 'Ingénieur', 'None', 'None', 'Fouilles archéologiques',
# 'du 6/08/2012 au 24/08/2012', 'None', 'M. Didier  VERMEERSCH')

# """ stages(pays , organisme , secteur , typeDeStage ,
# etudiant ,  promo, diplome , domaine,  appro , sujet , datesStages, mailTuteur, tuteur )"""
import unicodedata

def enleveUnicode(text):
    if(isinstance(text,basestring)):
        return unicodedata.normalize('NFKD', text).encode('ascii','ignore')
    else:
        return text
def faire(x):
    return map(enleveUnicode,list(x))

def regarder(texte):
    global data;
    cur.execute(texte)
    data=cur.fetchall()
    data=map(faire,data)
    for x in data:
        print x
    print data.__len__()


###############
# Print in xls

import xlsxwriter

def getLength(x):
    return x.__len__()

def enregistrer(nameFile):
    global data;
    dossier='/home/nicolas/Bureau/'
    workbook = \
        xlsxwriter.Workbook(dossier+nameFile)
    caption = 'Default table with data.'
    worksheet2 = workbook.add_worksheet()
    worksheet2.add_table(0,0,data.__len__()-1,
                         data[0].__len__()-1, {'data': data})
    data3=np.array(data)
    for i in range(data3.shape[1]):
        width = np.percentile(map(getLength,data3[:,i]),80.0)
        worksheet2.set_column(i, i, int(width));
    workbook.close()

##################
regarder("""SELECT pays, promo, etudiant, organisme, sujet FROM stages
                WHERE (promo >= 2014)
                        AND NOT (typeDeStage LIKE "%Decouverte%")
                        AND NOT (typeDeStage = "STAGE LONG")
                        AND  ((typeDeStage LIKE '%end%') OR (typeDeStage LIKE '%fin%'))
                        AND diplome LIKE '%Ingeni%'
                ORDER BY pays ASC, organisme ASC
                        """)
# AND NOT(pays LIKE "%FRANCE%")
enregistrer("recent.xlsx")

#promo, etudiant, pays, organisme, sujet FROM stages
regarder("""SELECT pays, promo, etudiant, organisme, sujet FROM stages
                WHERE (domaine LIKE "%Decision%")
                AND  ((typeDeStage LIKE '%end%') OR (typeDeStage LIKE '%fin%'))
                ORDER BY pays ASC, promo ASC""") #promo ASC, pays ASC
#AND NOT(pays LIKE "%FRANCE%")
enregistrer("sid.xlsx")

#promo, etudiant, pays, organisme, sujet FROM stages
regarder("""SELECT pays, promo, etudiant, organisme, sujet FROM stages
                WHERE (appro LIKE "%Autom%")
                AND  ((typeDeStage LIKE '%end%') OR (typeDeStage LIKE '%fin%'))
                ORDER BY pays ASC, promo ASC""") #promo ASC, pays ASC
#AND NOT(pays LIKE "%FRANCE%")
enregistrer("autom.xlsx")

regarder("""SELECT pays, promo, etudiant, organisme, sujet FROM stages
                WHERE (appro LIKE "%Image%")
                AND  ((typeDeStage LIKE '%end%') OR (typeDeStage LIKE '%fin%'))
                ORDER BY pays ASC, promo ASC""")
# AND NOT(pays LIKE "%FRANCE%")
enregistrer("si.xlsx")


regarder("""SELECT pays, promo, etudiant, organisme, sujet FROM stages
                WHERE pays IN ('ROYAUME UNI' , 'GRANDE BRETAGNE')
                AND  ((typeDeStage LIKE '%end%') OR (typeDeStage LIKE '%fin%'))
                ORDER BY pays ASC, promo ASC
""")
enregistrer("uk.xlsx")



























#####################
#essais




wb = px.Workbook()
for r in data:
    for c in r:
        ws.cell(row=r, column=c).value = data[]
    r += 1
wb.save('/home/nicolas/Bureau/machin.xlsx')


###########


import csv
csvfile = open('/home/nicolas/Bureau/machin.csv', 'wb')
truc = csv.writer(csvfile)
def enleveUnicode(text):
    return unicodedata.normalize('NFKD', text).encode('ascii','ignore')
def f(manyTexts):
    return map(g,manyTexts)
for x in data:
    truc.writerow(f(x))


###############"

title = u"Klüft skräms inför på fédéral électoral große"
import unicodedata
unicodedata.normalize('NFKD', title).encode('ascii','ignore')
'Kluft skrams infor pa federal electoral groe'

##################

for row in p.iter_rows():
    for k in row:
        a.append(k.value)



# convert list a to matrix (for example 5*6)
aa= np.resize(a, [5, 6])

# save matrix aa as xlsx file
WW=px.Workbook()
pp=WW.get_active_sheet()
pp.title='NEW_DATA'

f={'A':0,'B':1,'C':2,'D':3,'E':4,'F':5}

#insert values in six columns
for (i,j) in f.items():
    for k in np.arange(1,len(aa)+1):
        pp.cell('%s%d'%(i,k)).value=aa[k-1][j]

WW.save('newfilname.xlsx')


####################"

from openpyxl import load_workbook
wb2 = load_workbook('/home/nicolas/Bureau/job/data.xlsx')
print wb2.get_sheet_names()

aaa=wb2.get_active_sheet()



###########

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

#################

import csv, sqlite3

db="db1"
con = sqlite3.connect(db)
cur = con.cursor()
cur.execute("CREATE TABLE t (col1, col2);")

with open('data.csv','rb') as fin: # `with` statement available in 2.5+
    # csv.DictReader uses first line in file for column headings by default
    dr = csv.DictReader(fin) # comma is default delimiter
    to_db = [(i['col1'], i['col2']) for i in dr]

cur.executemany("INSERT INTO t (col1, col2) VALUES (?, ?);", to_db)
con.commit()



##########
# python csv2sqlite.py nameofdb directory

from __future__ import print_function
import sqlite3
import csv
import os
import glob
import sys

db = sys.argv[1]

conn = sqlite3.connect(db)
conn.text_factory = str  # allows utf-8 data to be stored

c = conn.cursor()

# traverse the directory and process each .csv file
for csvfile in glob.glob(os.path.join(sys.argv[2], "*.csv")):
    # remove the path and extension and use what's left as a table name
    tablename = os.path.splitext(os.path.basename(csvfile))[0]

    with open(csvfile, "rb") as f:
        reader = csv.reader(f)

        header = True
        for row in reader:
            if header:
                # gather column names from the first row of the csv
                header = False

                sql = "DROP TABLE IF EXISTS %s" % tablename
                c.execute(sql)
                sql = "CREATE TABLE %s (%s)" % (tablename,
                          ", ".join([ "%s text" % column for column in row ]))
                c.execute(sql)

                for column in row:
                    if column.lower().endswith("_id"):
                        index = "%s__%s" % ( tablename, column )
                        sql = "CREATE INDEX %s on %s (%s)" % ( index, tablename, column )
                        c.execute(sql)

                insertsql = "INSERT INTO %s VALUES (%s)" % (tablename,
                            ", ".join([ "?" for column in row ]))

                rowlen = len(row)
            else:
                # skip lines that don't have the right number of columns
                if len(row) == rowlen:
                    c.execute(insertsql, row)

        conn.commit()

c.close()
conn.close()

####################

import sqlite3 as lite
import os

personnes = [("Ronaldo", "Madrid"),
             ("Messi", "Barcelona"),
             ("Ibra", "Paris"),
             ("Rooney", "Manchester"),
             ("Casillas", "Madrid"),
             ("Bale", "Madrid")]

etats = [
    (1, "Ronaldo", 1, "party", 1),     # parc 1 appartient 'a Ronaldo depuis Janvier
    (2, "Messi", 2, "sleep", 2),
    (3, "Ibra", 2, "tatoo", 1),
    (4, "Rooney", 3, "party", 1),      # Mars
    (5, "Casillas", 3, "sleep", 1),
    (1, "Bale", 4, "coiffure", 1),     # Bale achete 1 'a Ronaldo
    (6, "Ronaldo", 4, "party", 1),     # Ronaldo achete 6
    (5, "Bale", 4, "coiffure", 1),     # Bale achete 5 'a Casillas
    (3, "Casillas", 5, "sleep", 1),    # Casillas achete 3 'a Ibra
    (1, "Ibra", 5, "tatoo", 1)         # Ibra achete 1 'a Bale
]

quartiers = [
    (1, 1, 1),
    (1, 2, 2),
    (2, -3, -3),
    (2, -4, -4),
    (2, -5, -5)
]

# P1: Ronaldo, Bale, Ibra   => 3
# P2: Messi                 => 1
# P3: Ibra, Casillas        => 2
# P4: Rooney                => 2
# P5: Caillas, Bale         => 2

# creer la base de donnes
try: os.remove("bdonnes.db")
except: pass
con = lite.connect("bdonnes.db")
cur = con.cursor()

# creer tableaux
cur.execute("CREATE TABLE LesEtatsParcelles(numP INT, nom TEXT, dateDepuis INT, sol TEXT, noEns INT)")
cur.execute("CREATE TABLE LesPersonnes(nom TEXT, adresse TEXT)")
cur.execute("CREATE TABLE LesEnsQuartiers(noEns INT, absc INT, ord INT)")

# inserer information dans les tableaux
for numP, nom, date, sol, noEns in etats:
    cur.execute("INSERT INTO LesEtatsParcelles VALUES(%d, '%s', %d, '%s', %d)" % (numP, nom, date, sol, noEns))

for nom, adresse in personnes:
    cur.execute("INSERT INTO LesPersonnes VALUES('%s', '%s')" % (nom, adresse))

for noEns, absc, orde in quartiers:
    cur.execute("INSERT INTO LesEnsQuartiers VALUES(%d, %d, %d)" % (noEns, absc, orde))


# requetes 1, 2, 3, 4

print "SQL Req. 1"

# marche pas
cur.execute("SELECT nom, sol FROM LesEtatsParcelles WHERE numP = 1") # pas de 15 dans l'exemple
print cur.fetchall()

# marche
cur.execute("SELECT nom, sol FROM LesEtatsParcelles WHERE numP = 1 AND dateDepuis = (SELECT MAX(dateDepuis) FROM LesEtatsParcelles WHERE numP = 1)")
print cur.fetchall()

print
print "SQL Req. 2"

# marche
cur.execute("SELECT LesPersonnes.nom FROM LesPersonnes INNER JOIN LesEtatsParcelles ON (LesPersonnes.nom = LesEtatsParcelles.nom) WHERE LesPersonnes.adresse = 'Madrid' AND LesEtatsParcelles.numP = 1") # pas de 15 dans l'exemple
print cur.fetchall()

print
print "SQL Req. 3"

# marche pas
cur.execute("SELECT numP AS 'numPero_de_parcelle', COUNT(DISTINCT(nom)) AS 'nombres_propres' FROM LesEtatsParcelles")
print cur.fetchall()

# marche
cur.execute("SELECT numP, COUNT(DISTINCT(nom)) FROM LesEtatsParcelles GROUP BY numP")
print cur.fetchall()

print
print "SQL Req. 4"

# marche
cur.execute("SELECT LesEnsQuartiers.absc, LesEnsQuartiers.ord FROM LesEnsQuartiers INNER JOIN LesEtatsParcelles ON LesEnsQuartiers.noEns=LesEtatsParcelles.noEns WHERE LesEtatsParcelles.numP=2")
print cur.fetchall()

# marche
cur.execute("SELECT q.absc, q.ord FROM LesEnsQuartiers q, LesEtatsParcelles p WHERE p.numP=2 AND p.noEns=q.noEns")
print cur.fetchall()

############################
